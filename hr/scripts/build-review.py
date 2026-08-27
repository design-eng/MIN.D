# -*- coding: utf-8 -*-
"""연봉조정 평가표 — 점수 입력 → 등급 → 인상률 → 조정 연봉 자동 산출"""
import os, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

F = '맑은 고딕'
INK, GRAY, BLUE, ACC = '14161B', '6B7280', '0000FF', 'C2410C'
SURF, YEL, BAND, SOFT = 'F4F5F7', 'FFF9DB', 'E8EAED', 'FAFBFC'
thin = Side(style='thin', color='D4D6DB')
med  = Side(style='medium', color='14161B')
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)

FIRST, LAST = 6, 15          # 직원 10명
wb = openpyxl.Workbook()

def title_block(ws, t, sub):
    ws['A1'] = t; ws['A1'].font = Font(name=F, size=16, bold=True, color=INK)
    ws['A2'] = sub; ws['A2'].font = Font(name=F, size=9, color=GRAY)
    ws.sheet_view.showGridLines = False

def hrow(ws, row, headers, widths, auto=()):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=F, size=9, bold=True, color=GRAY if i in auto else INK)
        c.fill = PatternFill('solid', fgColor=BAND if i in auto else SURF)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = Border(left=thin, right=thin, top=med, bottom=med)
        ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    ws.row_dimensions[row].height = 32

def cell(ws, r, c, v=None, fmt=None, auto=False, center=True, bold=False):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(name=F, size=9, bold=bold, color=INK if auto else BLUE)
    x.border = BOX
    x.alignment = Alignment(horizontal='center' if center else 'left', vertical='center')
    if auto: x.fill = PatternFill('solid', fgColor=SOFT)
    if fmt: x.number_format = fmt
    return x

# ═════════ 1. 사용안내 ═════════
ws = wb.active; ws.title = '사용안내'
title_block(ws, '연봉조정 평가표', '취업규칙 제34조 · 연봉조정 운영기준에 따른 산출 도구')
ws.column_dimensions['A'].width = 15; ws.column_dimensions['B'].width = 92
rows = [
    ('', ''),
    ('쓰는 순서', ''),
    ('1', '「최저임금」 시트에 해당 연도 최저임금 시급을 넣습니다. 매년 바뀝니다.'),
    ('2', '「평가표」에 직원 이름과 현재 연봉을 넣습니다.'),
    ('3', '자기평가와 회사평가 점수를 각 항목에 넣으면 총점·등급·권장 인상률이 자동 계산됩니다.'),
    ('4', '면담 후 「적용 인상률」을 넣으면 조정 연봉과 월 지급액이 나옵니다.'),
    ('5', '「검증」 열이 확인 필요로 뜨면 최저임금 미달이므로 금액을 올려야 합니다.'),
    ('6', '합의가 끝나면 「연봉이력」에 한 줄 남기고, 연봉계약서를 작성합니다.'),
    ('', ''),
    ('꼭 알아두실 점', ''),
    ('감액 불가', '연봉은 근로자 동의 없이 낮출 수 없습니다. D등급이어도 동결이 하한입니다. (취업규칙 제34조제5항)'),
    ('합의 결렬', '합의가 안 되면 종전 연봉이 그대로 유지됩니다. 이를 이유로 불이익을 주면 안 됩니다. (제34조제6항)'),
    ('시기', '입사일 기준입니다. 입사 기념일이 속한 달의 다음 달 1일부터 적용하고, 면담은 그 전달까지 마칩니다.'),
    ('권장 구간', '등급별 인상률은 권장 범위입니다. 경영 상황에 따라 구간 안에서 정하되, 구간을 벗어나면 면담에서 사유를 설명하세요.'),
    ('반영 절차', '연봉이 바뀌면 4대보험 보수월액 변경을 신고하고, 근태·연차 관리대장의 연봉도 갱신해야 합니다.'),
    ('평가 금지', '휴가 사용, 출산·육아휴직, 성별·연령 등 차별 사유, 괴롭힘 신고는 평가에 반영할 수 없습니다.'),
]
r = 4
for k, v in rows:
    a = ws.cell(row=r, column=1, value=k); b = ws.cell(row=r, column=2, value=v)
    head = k and not v
    a.font = Font(name=F, size=10 if head else 9, bold=True, color=INK if head else GRAY)
    b.font = Font(name=F, size=9, color=INK)
    a.alignment = Alignment(vertical='top'); b.alignment = Alignment(vertical='top', wrap_text=True)
    if head:
        for col in (1, 2):
            ws.cell(row=r, column=col).border = Border(bottom=Side(style='thin', color=INK))
        ws.row_dimensions[r].height = 24
    elif v and len(v) > 68: ws.row_dimensions[r].height = 28
    r += 1

# ═════════ 2. 최저임금 ═════════
mw = wb.create_sheet('최저임금')
title_block(mw, '최저임금 기준', '해당 연도 시급만 넣으면 나머지는 자동 계산됩니다. 고용노동부 고시로 매년 바뀝니다.')
for col, w in zip('AB', (26, 20)): mw.column_dimensions[col].width = w
items = [
    ('최저임금 시급 (원)', 10320, True,  '#,##0'),
    ('월 소정근로시간',    209,   True,  '0'),
    ('월 환산액 (원)',     '=B4*B5', False, '#,##0'),
    ('수습 감액 하한 90% (원)', '=ROUND(B6*0.9,0)', False, '#,##0'),
]
mw['A3'] = '항목'; mw['B3'] = '값'
for c in ('A3','B3'):
    mw[c].font = Font(name=F, size=9, bold=True, color=INK)
    mw[c].fill = PatternFill('solid', fgColor=SURF)
    mw[c].alignment = Alignment(horizontal='center')
    mw[c].border = Border(left=thin, right=thin, top=med, bottom=med)
for i, (lbl, val, inp, fmt) in enumerate(items, start=4):
    a = mw.cell(row=i, column=1, value=lbl)
    a.font = Font(name=F, size=9, color=INK); a.border = BOX
    a.alignment = Alignment(vertical='center')
    b = mw.cell(row=i, column=2, value=val)
    b.font = Font(name=F, size=9, bold=not inp, color=BLUE if inp else INK)
    b.border = BOX; b.number_format = fmt
    b.alignment = Alignment(horizontal='center', vertical='center')
    if inp: b.fill = PatternFill('solid', fgColor=YEL)
    else:   b.fill = PatternFill('solid', fgColor=SOFT)
mw['A9'] = '※ 2026년 시급 10,320원을 기본값으로 넣어 두었습니다. 매년 고시를 확인해 B4를 고치세요.'
mw['A9'].font = Font(name=F, size=9, color=ACC)

# ═════════ 3. 평가기준 ═════════
st = wb.create_sheet('평가기준')
title_block(st, '평가 항목과 등급', '연봉조정 운영기준 2항·3항. 배점을 바꾸면 평가표의 만점도 함께 고쳐야 합니다.')
hrow(st, 4, ['영역', '세부 항목', '배점'], [16, 46, 8])
AREAS = [
    ('직무 성과', '산출물의 완성도와 품질', 15),
    ('', '일정 준수와 진행 관리', 15),
    ('', '클라이언트 요구 파악과 대응', 10),
    ('직무 역량', '사용자 조사와 화면 설계의 깊이', 10),
    ('', '디자인 시스템 구축·운용', 10),
    ('', '도구 숙련도와 새로운 방법의 습득', 10),
    ('협업과 태도', '팀·직무 간 소통과 정보 공유', 10),
    ('', '피드백 수용과 반영', 10),
    ('기여와 성장', '업무 방식·프로세스 개선 제안', 5),
    ('', '학습 내용의 공유와 후배 지원', 5),
]
for i, (a, b, p) in enumerate(AREAS, start=5):
    for col, v, ctr in ((1, a, True), (2, b, False), (3, p, True)):
        x = st.cell(row=i, column=col, value=v)
        x.font = Font(name=F, size=9, color=INK, bold=(col == 1 and bool(v)))
        x.border = BOX
        x.alignment = Alignment(horizontal='center' if ctr else 'left', vertical='center')
st.cell(row=15, column=2, value='합계').font = Font(name=F, size=9, bold=True, color=INK)
st.cell(row=15, column=3, value='=SUM(C5:C14)').font = Font(name=F, size=9, bold=True, color=INK)
for col in (1, 2, 3):
    st.cell(row=15, column=col).border = Border(top=med, bottom=med, left=thin, right=thin)
    st.cell(row=15, column=col).fill = PatternFill('solid', fgColor=SURF)
    st.cell(row=15, column=col).alignment = Alignment(horizontal='center')

st['A18'] = '등급과 인상률'
st['A18'].font = Font(name=F, size=11, bold=True, color=INK)
hrow(st, 19, ['등급', '총점', '정의', '인상률'], [16, 46, 8, 8])
st.cell(row=19, column=2, value='총점'); st.cell(row=19, column=3, value='정의')
GRADES = [('S', '90 ~ 100', '기대를 크게 넘어섰다', '8 ~ 12%'),
          ('A', '80 ~ 89',  '기대를 넘어섰다',      '5 ~ 8%'),
          ('B', '65 ~ 79',  '기대한 수준을 충족했다', '3 ~ 5%'),
          ('C', '50 ~ 64',  '일부 항목에서 보완이 필요하다', '0 ~ 3%'),
          ('D', '49 이하',  '전반적인 개선이 필요하다', '동결')]
for i, row in enumerate(GRADES, start=20):
    for col, v in enumerate(row, start=1):
        x = st.cell(row=i, column=col, value=v)
        x.font = Font(name=F, size=9, color=INK, bold=(col == 1))
        x.border = BOX
        x.alignment = Alignment(horizontal='center' if col != 3 else 'left', vertical='center')
st['A26'] = '※ D등급이라도 연봉을 감액할 수 없습니다. 동결이 하한입니다.'
st['A26'].font = Font(name=F, size=9, color=ACC)

# ═════════ 4. 평가표 ═════════
ev = wb.create_sheet('평가표')
title_block(ev, '평가표', '자기평가와 회사평가를 각각 넣으면 총점·등급·권장 인상률이 자동 계산됩니다.')
ev['A3'] = '입력 방법 :  파란 글씨 = 직접 입력    ·    검은 글씨 + 회색 배경 = 자동 계산 (수정 금지)'
ev['A3'].font = Font(name=F, size=9, bold=True, color=ACC)
ev['J3'] = '점수는 평가기준 시트의 배점 이내로 넣습니다. 총점은 회사평가 기준으로 산정합니다.'
ev['J3'].font = Font(name=F, size=9, color=GRAY)
hdr = ['성명', '입사일', '현 연봉', '자기평가\n총점', '성과\n40', '역량\n30', '협업\n20', '기여\n10',
       '회사평가\n총점', '등급', '권장\n인상률', '적용\n인상률', '조정 연봉', '월 지급액', '증감액', '검증', '적용일']
hrow(ev, 5, hdr, [11, 12, 13, 10, 8, 8, 8, 8, 10, 8, 11, 10, 14, 13, 12, 12, 13],
     auto=(9, 10, 11, 13, 14, 15, 16, 17))
ev.freeze_panes = 'B6'

for r in range(FIRST, LAST + 1):
    cell(ev, r, 1)                                  # 성명
    cell(ev, r, 2, fmt='yyyy-mm-dd')                # 입사일
    cell(ev, r, 3, fmt='#,##0')                     # 현 연봉
    cell(ev, r, 4, fmt='0')                         # 자기평가 총점
    for col in (5, 6, 7, 8): cell(ev, r, col, fmt='0')
    # 회사평가 총점
    cell(ev, r, 9, f'=IF($A{r}="","",SUM(E{r}:H{r}))', fmt='0', auto=True, bold=True)
    # 등급
    cell(ev, r, 10, f'=IF($A{r}="","",IF(I{r}>=90,"S",IF(I{r}>=80,"A",IF(I{r}>=65,"B",'
                    f'IF(I{r}>=50,"C","D")))))', auto=True, bold=True)
    # 권장 인상률
    cell(ev, r, 11, f'=IF($A{r}="","",IF(J{r}="S","8 ~ 12%",IF(J{r}="A","5 ~ 8%",'
                    f'IF(J{r}="B","3 ~ 5%",IF(J{r}="C","0 ~ 3%","동결")))))', auto=True)
    # 적용 인상률 (입력)
    cell(ev, r, 12, fmt='0.0%')
    # 조정 연봉 — 천원 단위 반올림, 감액 금지
    cell(ev, r, 13, f'=IF(OR($A{r}="",C{r}=""),"",MAX(C{r},ROUND(C{r}*(1+N(L{r}))/1000,0)*1000))',
         fmt='#,##0', auto=True, bold=True)
    # 월 지급액
    cell(ev, r, 14, f'=IF(M{r}="","",ROUND(M{r}/12,0))', fmt='#,##0', auto=True)
    # 증감액
    cell(ev, r, 15, f'=IF(M{r}="","",M{r}-C{r})', fmt='#,##0', auto=True)
    # 최저임금 검증
    cell(ev, r, 16, f'=IF(N{r}="","",IF(N{r}>=최저임금!$B$6,"적합","확인 필요"))', auto=True, bold=True)
    # 적용일 — 입사 기념일이 속한 달의 다음 달 1일
    cell(ev, r, 17, f'=IF(B{r}="","",DATE(YEAR($B$2)+IF(MONTH(B{r})>=MONTH($B$2),0,1),'
                    f'MONTH(B{r})+1,1))', fmt='yyyy-mm-dd', auto=True)

ev['A2'] = '평가 기준일'
ev['A2'].font = Font(name=F, size=9, bold=True, color=INK)
ev['B2'] = datetime.date(2026, 12, 1)
ev['B2'].font = Font(name=F, size=10, bold=True, color=BLUE)
ev['B2'].number_format = 'yyyy-mm-dd'
ev['B2'].fill = PatternFill('solid', fgColor=YEL); ev['B2'].border = BOX
ev['C2'] = '← 이 날짜가 속한 회차를 기준으로 적용일을 계산합니다.'
ev['C2'].font = Font(name=F, size=9, color=GRAY)

# 예시 행
exr = FIRST
ev.cell(row=exr, column=1, value='강성아')
ev.cell(row=exr, column=2, value=datetime.date(2026, 1, 12)).number_format = 'yyyy-mm-dd'
ev.cell(row=exr, column=3, value=32000000).number_format = '#,##0'
for col, v in ((4, 78), (5, 32), (6, 24), (7, 17), (8, 8)):
    ev.cell(row=exr, column=col, value=v)
ev.cell(row=exr, column=12, value=0.04).number_format = '0.0%'
for col in (1, 2, 3, 4, 5, 6, 7, 8, 12):
    ev.cell(row=exr, column=col).font = Font(name=F, size=9, color=GRAY, italic=True)

n = LAST + 2
notes = [
 '· 「조정 연봉」은 천원 단위로 반올림하며, 어떤 경우에도 현 연봉보다 낮아지지 않습니다. (취업규칙 제34조제5항)',
 '· 「검증」이 확인 필요로 뜨면 월 지급액이 최저임금 월 환산액에 미달한다는 뜻입니다. 금액을 올려야 합니다.',
 '· 「적용일」은 입사 기념일이 속한 달의 다음 달 1일입니다. 면담은 그 전달까지 마칩니다.',
 '· 예시 행(6행)은 실제 입력 전에 지우세요.',
]
for i, t in enumerate(notes):
    ev.cell(row=n + i, column=1, value=t).font = Font(name=F, size=9, color=GRAY)

# ═════════ 5. 연봉이력 ═════════
hs = wb.create_sheet('연봉이력')
title_block(hs, '연봉이력', '조정이 확정될 때마다 한 줄씩 남깁니다. 퇴직금 산정과 다음 조정의 근거가 됩니다.')
hs['A3'] = '입력 방법 :  파란 글씨 = 직접 입력    ·    검은 글씨 + 회색 배경 = 자동 계산'
hs['A3'].font = Font(name=F, size=9, bold=True, color=ACC)
hrow(hs, 5, ['적용일', '성명', '종전 연봉', '조정 연봉', '증감액', '인상률', '등급', '비고'],
     [13, 11, 14, 14, 13, 10, 8, 34], auto=(5, 6))
hs.freeze_panes = 'A6'
for r in range(6, 106):
    cell(hs, r, 1, fmt='yyyy-mm-dd')
    cell(hs, r, 2)
    cell(hs, r, 3, fmt='#,##0')
    cell(hs, r, 4, fmt='#,##0')
    cell(hs, r, 5, f'=IF(OR(C{r}="",D{r}=""),"",D{r}-C{r})', fmt='#,##0', auto=True)
    cell(hs, r, 6, f'=IF(OR(C{r}="",D{r}="",C{r}=0),"",D{r}/C{r}-1)', fmt='0.0%', auto=True)
    cell(hs, r, 7)
    cell(hs, r, 8, center=False)
d = DataValidation(type='list', formula1='"S,A,B,C,D"', allow_blank=True, showDropDown=False)
hs.add_data_validation(d); d.add('G6:G105')

wb.calculation.fullCalcOnLoad = True
wb.save(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '마인드_연봉조정_평가표.xlsx'))
print('saved', [s.title for s in wb.worksheets])

# -*- coding: utf-8 -*-
"""빈 템플릿(build-workbook.py 산출물)에 현재 재직자를 입력하고 예시 행을 지운다.

관리대장에는 근태·연차 산정에 필요한 항목만 넣는다. 생년월일·연락처·주소 등은
근로계약서에 있고 이 대장의 용도와 무관하므로 옮기지 않는다.
"""
import os, datetime
import openpyxl
from openpyxl.styles import Font

F = '맑은 고딕'
BLUE = '0000FF'
HERE = os.path.dirname(os.path.abspath(__file__))
PATH = os.path.join(HERE, '..', '마인드_근태연차_관리대장.xlsx')

# 사번 · 성명 · 입사일 · 직위 · 고용형태 · 재직상태 · 퇴사일 · 연봉 · 주소정근로 · 비고
ROSTER = [
    ('M001', '강성아', datetime.date(2026, 1, 12), 'UI·UX 디자이너', '정규사원', '재직',
     None, 32000000, 40, '수습 2026-01-12 ~ 2026-04-11 종료'),
]

EMP_FIRST, EMP_LAST = 5, 24
LV_FIRST, AT_FIRST = 5, 5

wb = openpyxl.load_workbook(PATH)
em = wb['직원명부']

# 예시 행을 포함해 입력 칸을 전부 비운다 (수식 열 I 는 건드리지 않는다)
INPUT_COLS = [1, 2, 3, 4, 5, 6, 7, 8, 10, 11]
for r in range(EMP_FIRST, EMP_LAST + 1):
    for c in INPUT_COLS:
        em.cell(row=r, column=c).value = None

for i, row in enumerate(ROSTER, start=EMP_FIRST):
    for c, v in zip(INPUT_COLS, row):
        cell = em.cell(row=i, column=c, value=v)
        cell.font = Font(name=F, size=9, color=BLUE)   # 예시 행의 회색 이탤릭을 정상 입력 서식으로

# 휴가대장·근태기록의 예시 행 삭제 — 실제 기록은 회사가 채운다
for sheet, first, cols in (('휴가대장', LV_FIRST, [2, 4, 5, 6, 7, 8, 9]),
                           ('근태기록', AT_FIRST, [1, 3, 5, 6, 7, 10, 11])):
    ws = wb[sheet]
    for c in cols:
        ws.cell(row=first, column=c).value = None

wb.calculation.fullCalcOnLoad = True
wb.save(PATH)
print('입력 완료:', [r[1] for r in ROSTER])

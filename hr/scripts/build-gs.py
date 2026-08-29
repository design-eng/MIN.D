# -*- coding: utf-8 -*-
"""검증된 관리대장 xlsx 에서 수식·서식·데이터를 뽑아 Apps Script 를 생성한다.

손으로 옮기면 오타가 난다. 실제로 LibreOffice 재계산까지 마친 xlsx 를 원본으로
삼아 기계적으로 변환하므로 수식이 어긋날 여지가 없다.

산출물 : hr/scripts/gs-mind.gs  (구글 시트 Apps Script 에 통째로 붙여넣는 파일)
"""
import os, re, json, datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC  = os.path.join(HERE, '..', '마인드_근태연차_관리대장.xlsx')
PUNCH= os.path.join(HERE, 'gs-punch.gs')
OUT  = os.path.join(HERE, 'gs-mind.gs')

# 시트별 표 블록 : (머리글 행, 첫 데이터 행, 마지막 데이터 행)
# 보상휴가는 요약표와 사용기록표 두 덩어리로 되어 있다.
BODY = {'직원명부':[(4,5,24)], '연차현황':[(4,5,24)], '휴가대장':[(4,5,504)],
        '근태기록':[(4,5,2004)], '월별집계':[(4,5,24)], '프로젝트집계':[(4,5,34)],
        '보상휴가':[(4,5,14),(18,19,118)]}

def tmpl(f, first):
    """본보기 행의 상대참조만 {r} 자리표시자로 바꾼다.

    열 문자 바로 뒤에 붙은 행 번호만 잡으므로 $B$5 같은 절대참조와
    $H$2004 같은 범위 끝은 그대로 남는다.
    """
    rx = re.compile(r'([A-Z]{1,2})' + str(first) + r'(?![0-9])')
    return rx.sub(lambda m: m.group(1) + '{r}', f)

def jsval(v):
    if v is None: return None
    if isinstance(v, datetime.datetime):
        return {'__d': [v.year, v.month, v.day]}
    if isinstance(v, datetime.date):
        return {'__d': [v.year, v.month, v.day]}
    if isinstance(v, datetime.time):
        return {'__t': round((v.hour*3600 + v.minute*60 + v.second)/86400.0, 10)}
    if isinstance(v, bool): return v
    if isinstance(v, (int, float)): return v
    return str(v)

wb = openpyxl.load_workbook(SRC)
sheets = []

for ws in wb.worksheets:
    name = ws.title
    blocks = BODY.get(name, [])
    spec = {'name': name, 'freeze': ws.freeze_panes, 'widths': [],
            'blocks': [], 'statics': [], 'seed': []}

    maxc = ws.max_column
    for i in range(1, maxc + 1):
        L = openpyxl.utils.get_column_letter(i)
        d = ws.column_dimensions.get(L)
        spec['widths'].append(round((d.width or 9) * 7.2) if d else 68)

    # 블록별 열 정의 : 첫 데이터 행을 본보기로 삼는다
    for hdr, first, last in blocks:
        cols = []
        for i in range(1, maxc + 1):
            c = ws.cell(row=first, column=i)
            f = c.value if isinstance(c.value, str) and c.value.startswith('=') else None
            cols.append({'f': tmpl(f, first) if f else None,
                         'n': c.number_format if c.number_format != 'General' else None})
        spec['blocks'].append({'h': hdr, 'first': first, 'last': last, 'cols': cols})

    inBody = lambda r: any(f <= r <= l for _, f, l in blocks)

    # 표 밖의 고정 셀 (제목·안내·코드표·조회 기준일 등)
    for row in ws.iter_rows():
        for c in row:
            if c.value in (None, ''): continue
            if inBody(c.row): continue
            spec['statics'].append([c.row, c.column, jsval(c.value)])

    # 표 안에 이미 들어 있는 실제 기록 (수식이 아닌 것만)
    for _, first, last in blocks:
        for r in range(first, min(last, ws.max_row) + 1):
            for i in range(1, maxc + 1):
                v = ws.cell(row=r, column=i).value
                if v in (None, ''): continue
                if isinstance(v, str) and v.startswith('='): continue
                spec['seed'].append([r, i, jsval(v)])
    sheets.append(spec)

# 데이터 유효성 (드롭다운) — 코드표 참조
DV = [('직원명부', 'E', 5, 24, '코드표!$A$5:$A$8'),
      ('직원명부', 'F', 5, 24, '코드표!$B$5:$B$7'),
      ('휴가대장', 'E', 5, 504, '코드표!$D$5:$D$13'),
      ('근태기록', 'L', 5, 2004, '코드표!$C$5:$C$13')]

payload = json.dumps({'sheets': sheets, 'dv': DV}, ensure_ascii=False, separators=(',', ':'))

punch = open(PUNCH, encoding='utf-8').read()
# 기존 초기설정()은 출퇴근 시트만 만든다 → 이름을 바꿔 대장 생성 뒤에 부르게 한다
punch = punch.replace('function 초기설정() {', 'function 출퇴근시트만들기() {', 1)
# 대장만들기는 시트를 지우고 다시 만든다. 메뉴에 두면 실수로 눌러 기록이 날아간다.
punch = punch.replace("    .addSeparator()\n    .addItem('초기설정', '초기설정')\n", '')
punch = punch.replace(
  "  SpreadsheetApp.getUi().alert('초기설정이 끝났습니다.\\n\\n「출퇴근」 시트에서 체크박스를 눌러 보세요.');\n", '')
# 상단 주석/상수 블록은 중복되므로 본문만 남긴다
punch = punch[punch.index('const TZ ='):]

header = '''/**
 * 마인드 근태관리 — 구글 시트 전체를 만들고 운영하는 스크립트
 *
 * 쓰는 법
 *   1. sheets.google.com 에서 「빈 스프레드시트」를 하나 만든다
 *   2. 그 시트에서  확장 프로그램 → Apps Script
 *   3. Code.gs 내용을 전부 지우고 이 파일을 통째로 붙여넣는다
 *   4. 저장한 뒤 함수 목록에서 「대장만들기」를 골라 실행
 *
 * 엑셀 파일을 가져올 필요가 없다. 시트 9개와 수식·드롭다운·기존 기록까지
 * 스크립트가 직접 만든다. 다시 실행하면 처음부터 새로 만든다.
 *
 * 이 파일은 build-gs.py 가 검증된 관리대장 xlsx 에서 뽑아 생성한 것이다.
 * 수식을 고칠 때는 이 파일이 아니라 build-workbook.py 를 고치고 다시 뽑는다.
 */

'''

body = '''
var SPEC = %s;

var HEAD_FILL='#F4F5F7', AUTO_FILL='#E8EAED', LINE='#D4D6DB';
var INK='#14161B', GRAY='#6B7280', BLUE='#0000FF';

function 대장만들기(){
  var ss=SpreadsheetApp.getActive();
  ss.setSpreadsheetTimeZone(TZ);
  var made=[];
  SPEC.sheets.forEach(function(s){ buildSheet_(ss,s); made.push(s.name); });
  applyDV_(ss);
  출퇴근시트만들기();
  // 스크립트가 만든 시트 외의 기본 시트를 지운다
  ss.getSheets().forEach(function(sh){
    if(made.indexOf(sh.getName())<0 && sh.getName()!==SH_PUNCH){
      try{ ss.deleteSheet(sh); }catch(e){}
    }
  });
  ss.setActiveSheet(ss.getSheetByName(SH_PUNCH));
  SpreadsheetApp.getUi().alert(
    '대장을 만들었습니다.\\n\\n시트 '+(made.length+1)+'개가 생성되었습니다.\\n'+
    '「출퇴근」 시트에서 체크박스를 눌러 보세요.');
}

function toVal_(v){
  if(v && typeof v==='object'){
    if(v.__d) return new Date(v.__d[0], v.__d[1]-1, v.__d[2]);
    if(v.__t!==undefined) return v.__t;
  }
  return v;
}

function buildSheet_(ss,s){
  var sh=ss.getSheetByName(s.name);
  if(sh) ss.deleteSheet(sh);
  sh=ss.insertSheet(s.name);
  sh.setHiddenGridlines(true);

  s.widths.forEach(function(w,i){ sh.setColumnWidth(i+1, Math.max(40,w)); });

  // 고정 셀
  s.statics.forEach(function(t){ sh.getRange(t[0],t[1]).setValue(toVal_(t[2])); });

  // 제목 줄 서식
  sh.getRange(1,1).setFontSize(15).setFontWeight('bold').setFontColor(INK);
  sh.getRange(2,1).setFontSize(9).setFontColor(GRAY);

  s.blocks.forEach(function(b){
    var n=b.last-b.first+1, ncol=b.cols.length;
    var hdr=sh.getRange(b.h,1,1,ncol);
    hdr.setFontSize(9).setFontWeight('bold').setVerticalAlignment('middle')
       .setHorizontalAlignment('center').setWrap(true)
       .setBorder(true,true,true,true,true,true,LINE,SpreadsheetApp.BorderStyle.SOLID);
    sh.setRowHeight(b.h,34);

    b.cols.forEach(function(col,i){
      var rng=sh.getRange(b.first,i+1,n,1);
      if(col.f){
        var out=[];
        for(var r=b.first;r<=b.last;r++) out.push([col.f.replace(/\\{r\\}/g, r)]);
        rng.setFormulas(out);
        rng.setFontColor(INK).setBackground('#FAFBFC');
        sh.getRange(b.h,i+1).setBackground(AUTO_FILL).setFontColor(GRAY);
      } else {
        rng.setFontColor(BLUE);
        sh.getRange(b.h,i+1).setBackground(HEAD_FILL).setFontColor(INK);
      }
      if(col.n) rng.setNumberFormat(col.n);
      rng.setFontSize(9).setHorizontalAlignment(i===0?'left':'center');
    });
  });

  // 실제 기록
  s.seed.forEach(function(t){
    sh.getRange(t[0],t[1]).setValue(toVal_(t[2])).setFontColor(BLUE);
  });

  if(s.freeze){
    var m=/^([A-Z]+)(\\d+)$/.exec(s.freeze);
    if(m){
      var cols=0, L=m[1];
      for(var k=0;k<L.length;k++) cols=cols*26+(L.charCodeAt(k)-64);
      sh.setFrozenRows(+m[2]-1); sh.setFrozenColumns(cols-1);
    }
  }
  finish_(sh);
}

function finish_(sh){
  sh.getRange(1,1,sh.getMaxRows(),sh.getMaxColumns()).setFontFamily('Arial');
}

function applyDV_(ss){
  SPEC.dv.forEach(function(d){
    var sh=ss.getSheetByName(d[0]); if(!sh) return;
    var src=ss.getRange(d[4]);
    var rule=SpreadsheetApp.newDataValidation()
      .requireValueInRange(src,true).setAllowInvalid(true).build();
    sh.getRange(d[2], colNum_(d[1]), d[3]-d[2]+1, 1).setDataValidation(rule);
  });
}
function colNum_(L){ var n=0; for(var i=0;i<L.length;i++) n=n*26+(L.charCodeAt(i)-64); return n; }

''' % payload

open(OUT, 'w', encoding='utf-8').write(header + body + punch)
print('생성:', OUT, os.path.getsize(OUT), 'bytes')
print('시트:', [s['name'] for s in sheets])
print('수식 열 수:', sum(1 for s in sheets for b in s['blocks'] for c in b['cols'] if c['f']))
print('고정 셀:', sum(len(s['statics']) for s in sheets),
      '· 기록 셀:', sum(len(s['seed']) for s in sheets))

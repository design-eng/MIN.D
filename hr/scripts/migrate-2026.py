# -*- coding: utf-8 -*-
"""구버전(7시트) 관리대장에 쌓인 2026년 기록을 현행 9시트 대장으로 옮긴다.

구버전에는 야간시간·보상휴가·프로젝트 열이 없었으므로 열 번호가 다르다.
    구 J 근태구분 → 신 L,  구 K 비고 → 신 N
나머지(날짜·사번·출근·퇴근·휴게)는 열 위치가 같다.

이관과 함께 연말까지의 날짜·요일·사번·공휴일을 미리 채워 둔다.
매일 입력하는 칸이 출근·퇴근 두 개로 줄어든다.

사용법:  python3 migrate-2026.py <구버전파일>
"""
import os, sys, datetime
import openpyxl
from openpyxl.styles import Font

F, BLUE, GRAY = '맑은 고딕', '0000FF', '6B7280'
HERE = os.path.dirname(os.path.abspath(__file__))
DEST = os.path.join(HERE, '..', '마인드_근태연차_관리대장.xlsx')
SRC = sys.argv[1]

EMP_NO = 'M001'
AT_FIRST, AT_LAST = 5, 2004
FILL_TO = datetime.date(2026, 12, 31)

# 2026년 하반기 관공서 공휴일. 토요일과 겹치는 광복절·개천절은 대체공휴일이 생기고,
# 추석 연휴가 토요일과 겹치는 경우에는 대체공휴일이 생기지 않는다.
HOLIDAYS = {
    datetime.date(2026, 9, 24): '추석 연휴',
    datetime.date(2026, 9, 25): '추석',
    datetime.date(2026, 9, 26): '추석 연휴',
    datetime.date(2026, 10, 3): '개천절',
    datetime.date(2026, 10, 5): '개천절 대체공휴일',
    datetime.date(2026, 10, 9): '한글날',
    datetime.date(2026, 12, 25): '성탄절',
}
# 구버전에 빠져 있던 것 — 8월 15일이 토요일이라 8월 17일이 대체공휴일이다.
ANNOTATE = {datetime.date(2026, 8, 17): '광복절 대체공휴일'}


def ink(cell, value, blue=True, fmt=None):
    cell.value = value
    cell.font = Font(name=F, size=9, color=BLUE if blue else GRAY)
    if fmt:
        cell.number_format = fmt


src = openpyxl.load_workbook(SRC, data_only=False)
dst = openpyxl.load_workbook(DEST)
so, do = src['근태기록'], dst['근태기록']

# ── 1. 구버전 근태기록 이관 ──────────────────────────────
moved, last_date, row = 0, None, AT_FIRST
for r in range(5, so.max_row + 1):
    d = so.cell(row=r, column=1).value
    if not isinstance(d, datetime.datetime):
        continue                                   # 표 아래 주석 줄
    day = d.date()
    ink(do.cell(row=row, column=1), day, fmt='yyyy-mm-dd')
    ink(do.cell(row=row, column=3), so.cell(row=r, column=3).value)
    for old_c, new_c, fmt in ((5, 5, 'hh:mm'), (6, 6, 'hh:mm'), (7, 7, '0.0')):
        v = so.cell(row=r, column=old_c).value
        if v not in (None, ''):
            ink(do.cell(row=row, column=new_c), v, fmt=fmt)
    gubun = so.cell(row=r, column=10).value        # 구 J → 신 L
    if gubun:
        ink(do.cell(row=row, column=12), gubun)
    note = ANNOTATE.get(day) or so.cell(row=r, column=11).value   # 구 K → 신 N
    if note:
        ink(do.cell(row=row, column=14), note, blue=False)
    last_date, row, moved = day, row + 1, moved + 1

# ── 2. 남은 날짜 미리 채우기 ─────────────────────────────
added = 0
day = last_date + datetime.timedelta(days=1)
while day <= FILL_TO and row <= AT_LAST:
    ink(do.cell(row=row, column=1), day, fmt='yyyy-mm-dd')
    ink(do.cell(row=row, column=3), EMP_NO)
    note = HOLIDAYS.get(day) or ('주말' if day.weekday() >= 5 else None)
    if note:
        ink(do.cell(row=row, column=14), note, blue=False)
    day, row, added = day + datetime.timedelta(days=1), row + 1, added + 1

# ── 3. 휴가대장 예시 행 제거 ─────────────────────────────
# 구버전에 남아 있던 2026-05-04 연차 한 줄은 템플릿 예시이지 실제 사용분이 아니다.
lg = dst['휴가대장']
for c in (2, 4, 5, 6, 7, 8, 9):
    lg.cell(row=5, column=c).value = None

# ── 4. 조회 기준일을 시행일에 맞춘다 ──────────────────────
# 취업규칙과 보상휴가제 서면합의의 시행일이 2026년 9월 1일이다.
dst['월별집계']['B3'] = datetime.date(2026, 9, 1)
dst['보상휴가']['B3'] = datetime.date(2026, 9, 1)

dst.calculation.fullCalcOnLoad = True
dst.save(DEST)
print(f'이관 {moved}행 ({last_date}까지) · 사전입력 {added}행 · 마지막 행 {row - 1}')

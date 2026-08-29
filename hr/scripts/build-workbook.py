# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

F = '맑은 고딕'
INK, GRAY, BLUE, ACC = '14161B', '6B7280', '0000FF', 'C2410C'
SURF, YEL, BAND, SOFT_FILL = 'F4F5F7', 'FFF9DB', 'E8EAED', 'FAFBFC'

thin = Side(style='thin', color='D4D6DB')
med  = Side(style='medium', color='14161B')
BOX  = Border(left=thin, right=thin, top=thin, bottom=thin)

EMP_FIRST, EMP_LAST = 5, 24          # 직원 20명
LV_FIRST,  LV_LAST  = 5, 504         # 휴가 500건
AT_FIRST,  AT_LAST  = 5, 2004        # 근태 2000건

wb = openpyxl.Workbook()

def title_block(ws, title, sub):
    ws['A1'] = title
    ws['A1'].font = Font(name=F, size=16, bold=True, color=INK)
    ws['A2'] = sub
    ws['A2'].font = Font(name=F, size=9, color=GRAY)

def header_row(ws, row, headers, widths, auto_cols=()):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name=F, size=9, bold=True,
                      color=GRAY if i in auto_cols else INK)
        c.fill = PatternFill('solid', fgColor=BAND if i in auto_cols else SURF)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = Border(left=thin, right=thin, top=med, bottom=med)
        ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
    ws.row_dimensions[row].height = 30

def body_cell(ws, r, c, value=None, fmt=None, auto=False, center=True):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = Font(name=F, size=9, color=INK if auto else BLUE)
    cell.border = BOX
    cell.alignment = Alignment(horizontal='center' if center else 'left', vertical='center')
    if auto:
        cell.fill = PatternFill('solid', fgColor='FAFBFC')
    if fmt:
        cell.number_format = fmt
    return cell

def legend(ws, row, extra=None):
    ws.cell(row=row, column=1,
            value='입력 방법 :  파란 글씨 = 직접 입력하는 칸    ·    검은 글씨 + 회색 배경 = 자동 계산 (수정 금지)')
    ws.cell(row=row, column=1).font = Font(name=F, size=9, color=ACC, bold=True)
    if extra:
        # 헤더 행에 덮이지 않도록 같은 행의 오른쪽 칸에 이어 쓴다
        ws.cell(row=row, column=5, value=extra)
        ws.cell(row=row, column=5).font = Font(name=F, size=9, color=GRAY)

# ============================================================
# 1. 사용안내
# ============================================================
ws = wb.active
ws.title = '사용안내'
ws.sheet_view.showGridLines = False
title_block(ws, '마인드 근태 · 연차 관리대장', '취업규칙 제11조(출근 및 퇴근) · 제22조(연차유급휴가) 운영용')
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 96

guide = [
    ('', ''),
    ('시트 구성', ''),
    ('직원명부', '가장 먼저 채웁니다. 사번·성명·입사일을 넣으면 나머지 시트가 이 값을 자동으로 끌어옵니다.'),
    ('연차현황', '입사일만 있으면 발생·사용·잔여 일수가 전부 자동 계산됩니다. 입력할 곳은 「이월」 칸뿐입니다.'),
    ('휴가대장', '휴가를 쓸 때마다 한 줄씩 기록합니다. 이 기록이 연차현황의 「사용」 일수가 됩니다.'),
    ('근태기록', '출근·퇴근 시각을 매일 기록합니다. 근무시간과 연장시간은 자동으로 계산됩니다.'),
    ('월별집계', '조회할 연월 하나만 바꾸면 그 달의 근무일수·지각·결근·연차사용이 직원별로 집계됩니다.'),
    ('프로젝트집계', '근태기록의 「프로젝트」 칸을 모아 프로젝트별 투입시간을 집계합니다. 견적의 근거가 됩니다.'),
    ('보상휴가', '연장·야간·휴일근로 수당을 갈음하는 휴가의 발생·사용·잔여를 관리합니다.'),
    ('코드표', '드롭다운 목록의 원본입니다. 항목을 늘리려면 여기에 추가합니다.'),
    ('', ''),
    ('연차 계산 기준', ''),
    ('산정 기준일', '입사일 기준입니다. 사람마다 연차 발생일이 다릅니다. (취업규칙 제22조제3항)'),
    ('1년 미만', '1개월 개근할 때마다 1일씩, 최대 11일. 입사 1주년 전날까지 써야 하고 남으면 소멸합니다.'),
    ('1년 이상', '15일. 3년 이상 근속부터 2년마다 1일씩 늘어나며 최대 25일입니다.'),
    ('반차', '휴가대장의 「일수」 칸에 0.5를 넣으면 반차로 집계됩니다.'),
    ('미사용분', '산정기간이 끝나면 소멸하며, 소멸한 일수만큼 미사용 수당을 지급합니다. (취업규칙 제23조제4항)'),
    ('', ''),
    ('꼭 알아두실 점', ''),
    ('월차는 없습니다', '월차휴가는 2003년 근로기준법 개정으로 폐지된 제도입니다. 지금 「월차」라고 부르는 것은 '
                    '입사 1년 미만 직원에게 매월 1일씩 생기는 연차(최대 11일)를 가리키며, 이 대장에서는 '
                    '연차현황 시트에 함께 계산됩니다.'),
    ('5인 미만', '연차유급휴가는 상시 근로자 5명 미만 사업장에 법적으로 강제되지 않습니다. 마인드는 '
              '취업규칙 제22조로 이를 부여하기로 정했으므로, 그 시점부터 회사의 의무가 됩니다.'),
    ('보존 의무', '근태기록과 임금 관련 서류는 3년간 보존해야 합니다. (근로기준법 제42조)'),
    ('5인이 되면', '상시 5명 이상이 되는 순간 연차유급휴가·연차사용촉진·주 12시간 연장 한도가 법적 의무로 '
                '전환됩니다. 가산수당은 이미 취업규칙으로 지급하고 있습니다.'),
]
r = 4
for k, v in guide:
    a = ws.cell(row=r, column=1, value=k)
    b = ws.cell(row=r, column=2, value=v)
    is_head = k and not v
    a.font = Font(name=F, size=10 if is_head else 9,
                  bold=True, color=INK if is_head else GRAY)
    b.font = Font(name=F, size=9, color=INK)
    b.alignment = Alignment(vertical='top', wrap_text=True)
    a.alignment = Alignment(vertical='top')
    if is_head:
        for col in (1, 2):
            ws.cell(row=r, column=col).border = Border(bottom=Side(style='thin', color=INK))
        ws.row_dimensions[r].height = 24
    elif v and len(v) > 70:
        ws.row_dimensions[r].height = 30
    r += 1

ws.cell(row=r + 1, column=1, value='※')
ws.cell(row=r + 1, column=2,
        value='이 대장은 취업규칙 초안을 전제로 만든 관리 도구입니다. 취업규칙의 연차 조항을 바꾸면 '
              '연차현황 시트의 계산식(H열)도 함께 고쳐야 합니다.')
ws.cell(row=r + 1, column=1).font = Font(name=F, size=9, color=ACC)
ws.cell(row=r + 1, column=1).alignment = Alignment(vertical='top')
ws.cell(row=r + 1, column=2).font = Font(name=F, size=9, color=ACC)
ws.cell(row=r + 1, column=2).alignment = Alignment(vertical='top', wrap_text=True)
ws.row_dimensions[r + 1].height = 30

# ============================================================
# 7. 코드표  (먼저 만들어 두어야 데이터 유효성 참조 가능)
# ============================================================
cs = wb.create_sheet('코드표')
cs.sheet_view.showGridLines = False
title_block(cs, '코드표', '드롭다운 목록의 원본입니다. 항목을 추가하면 각 시트의 드롭다운에 반영됩니다.')

CODES = {
    'A': ('고용형태', ['정규사원', '계약사원', '단시간사원', '수습']),
    'B': ('재직상태', ['재직', '휴직', '퇴사']),
    'C': ('근태구분', ['정상', '지각', '조퇴', '외출', '결근', '휴가', '재택', '출장', '휴일근로']),
    'D': ('휴가종류', ['연차', '경조휴가', '병가', '공가', '출산전후휴가', '배우자출산휴가',
                   '육아휴직', '가족돌봄휴가', '무급휴가']),
}
for col, (name, items) in CODES.items():
    c = cs[f'{col}4']
    c.value = name
    c.font = Font(name=F, size=9, bold=True, color=INK)
    c.fill = PatternFill('solid', fgColor=SURF)
    c.alignment = Alignment(horizontal='center')
    c.border = Border(left=thin, right=thin, top=med, bottom=med)
    cs.column_dimensions[col].width = 16
    for i, it in enumerate(items, start=5):
        v = cs[f'{col}{i}']
        v.value = it
        v.font = Font(name=F, size=9, color=INK)
        v.border = BOX
        v.alignment = Alignment(horizontal='center')

def dv(sheet, col_letter, first, last, code_col, n_items):
    d = DataValidation(type='list',
                       formula1=f'=코드표!${code_col}$5:${code_col}${4+n_items}',
                       allow_blank=True, showDropDown=False)
    sheet.add_data_validation(d)
    d.add(f'{col_letter}{first}:{col_letter}{last}')

# ============================================================
# 2. 직원명부
# ============================================================
em = wb.create_sheet('직원명부')
em.sheet_view.showGridLines = False
title_block(em, '직원명부', '가장 먼저 채우는 시트입니다. 사번은 다른 시트에서 사람을 찾는 열쇠이므로 중복되면 안 됩니다.')
legend(em, 3)
hdr = ['사번', '성명', '입사일', '직위', '고용형태', '재직상태', '퇴사일',
       '연봉 (원)', '월 지급액', '주 소정근로시간', '비고']
header_row(em, 4, hdr, [10, 12, 13, 12, 12, 11, 13, 15, 15, 14, 26], auto_cols=(9,))
em.freeze_panes = 'A5'

for r in range(EMP_FIRST, EMP_LAST + 1):
    body_cell(em, r, 1)                       # 사번
    body_cell(em, r, 2)                       # 성명
    body_cell(em, r, 3, fmt='yyyy-mm-dd')     # 입사일
    body_cell(em, r, 4)                       # 직위
    body_cell(em, r, 5)                       # 고용형태
    body_cell(em, r, 6)                       # 재직상태
    body_cell(em, r, 7, fmt='yyyy-mm-dd')     # 퇴사일
    body_cell(em, r, 8, fmt='#,##0')          # 연봉
    body_cell(em, r, 9, f'=IF(H{r}="","",ROUND(H{r}/12,0))', fmt='#,##0', auto=True)
    body_cell(em, r, 10, fmt='0')             # 주 소정근로시간
    body_cell(em, r, 11, center=False)        # 비고

# 예시 행
ex = ['M001', '홍길동', '2025-03-02', '디자이너', '정규사원', '재직', None, 42000000, None, 40,
      '← 예시 행입니다. 실제 입력 전에 지우세요.']
for i, v in enumerate(ex, start=1):
    if i == 9:
        continue
    cell = em.cell(row=EMP_FIRST, column=i)
    if i == 3 and v:
        import datetime
        cell.value = datetime.date(2025, 3, 2)
    elif v is not None:
        cell.value = v
    cell.font = Font(name=F, size=9, color=GRAY, italic=True)

dv(em, 'E', EMP_FIRST, EMP_LAST, 'A', 4)
dv(em, 'F', EMP_FIRST, EMP_LAST, 'B', 3)

# ============================================================
# 3. 연차현황
# ============================================================
lv = wb.create_sheet('연차현황')
lv.sheet_view.showGridLines = False
title_block(lv, '연차현황', '입사일 기준 자동 계산입니다. 직접 입력하는 칸은 「이월」 하나뿐입니다.')
lv['A3'] = '기준일'
lv['A3'].font = Font(name=F, size=9, bold=True, color=INK)
lv['B3'] = '=TODAY()'
lv['B3'].font = Font(name=F, size=10, bold=True, color=BLUE)
lv['B3'].number_format = 'yyyy-mm-dd'
lv['B3'].fill = PatternFill('solid', fgColor=YEL)
lv['B3'].border = BOX
lv['C3'] = '← 특정 시점 기준으로 보려면 이 칸에 날짜를 직접 넣으세요.'
lv['C3'].font = Font(name=F, size=9, color=GRAY)

hdr = ['사번', '성명', '입사일', '재직상태', '근속', '산정기간 시작', '산정기간 종료',
       '발생일수', '이월', '총 부여', '사용', '잔여', '소멸까지']
header_row(lv, 4, hdr, [10, 11, 13, 11, 14, 15, 15, 11, 9, 11, 9, 10, 12],
           auto_cols=(1, 2, 3, 4, 5, 6, 7, 8, 10, 11, 12, 13))
lv.freeze_panes = 'C5'

for r in range(EMP_FIRST, EMP_LAST + 1):
    e = f'직원명부!'
    body_cell(lv, r, 1, f'=IF({e}A{r}="","",{e}A{r})', auto=True)
    body_cell(lv, r, 2, f'=IF({e}B{r}="","",{e}B{r})', auto=True)
    body_cell(lv, r, 3, f'=IF({e}C{r}="","",{e}C{r})', fmt='yyyy-mm-dd', auto=True)
    body_cell(lv, r, 4, f'=IF({e}F{r}="","",{e}F{r})', auto=True)
    # 근속  "N년 M개월"
    body_cell(lv, r, 5,
        f'=IF(C{r}="","",DATEDIF(C{r},$B$3,"Y")&"년 "&DATEDIF(C{r},$B$3,"YM")&"개월")', auto=True)
    # 산정기간 시작
    body_cell(lv, r, 6,
        f'=IF(C{r}="","",IF(DATEDIF(C{r},$B$3,"Y")<1,C{r},EDATE(C{r},DATEDIF(C{r},$B$3,"Y")*12)))',
        fmt='yyyy-mm-dd', auto=True)
    # 산정기간 종료
    body_cell(lv, r, 7,
        f'=IF(C{r}="","",EDATE(C{r},(DATEDIF(C{r},$B$3,"Y")+1)*12)-1)',
        fmt='yyyy-mm-dd', auto=True)
    # 발생일수
    body_cell(lv, r, 8,
        f'=IF(C{r}="","",IF(DATEDIF(C{r},$B$3,"Y")<1,'
        f'MIN(11,DATEDIF(C{r},$B$3,"M")),'
        f'MIN(25,15+INT((DATEDIF(C{r},$B$3,"Y")-1)/2))))', fmt='0.0', auto=True)
    # 이월 (입력)
    body_cell(lv, r, 9, fmt='0.0')
    # 총 부여
    body_cell(lv, r, 10, f'=IF(C{r}="","",H{r}+N(I{r}))', fmt='0.0', auto=True)
    # 사용  (휴가대장에서 집계)
    body_cell(lv, r, 11,
        f'=IF(C{r}="","",SUMIFS(휴가대장!$F${LV_FIRST}:$F${LV_LAST},'
        f'휴가대장!$B${LV_FIRST}:$B${LV_LAST},$A{r},'
        f'휴가대장!$E${LV_FIRST}:$E${LV_LAST},"연차",'
        f'휴가대장!$D${LV_FIRST}:$D${LV_LAST},">="&$F{r},'
        f'휴가대장!$D${LV_FIRST}:$D${LV_LAST},"<="&$G{r}))', fmt='0.0', auto=True)
    # 잔여
    c = body_cell(lv, r, 12, f'=IF(C{r}="","",J{r}-K{r})', fmt='0.0', auto=True)
    c.font = Font(name=F, size=9, bold=True, color=INK)
    # 소멸까지
    body_cell(lv, r, 13,
        f'=IF(C{r}="","",IF(G{r}-$B$3<0,"기간종료",G{r}-$B$3&"일"))', auto=True)

nrow = EMP_LAST + 2
lv.cell(row=nrow, column=1,
        value='· 「발생일수」는 개근을 전제로 계산합니다. 결근이 있어 출근율 80% 미만인 직원은 수동으로 조정하세요.')
lv.cell(row=nrow + 1, column=1,
        value='· 「사용」은 휴가대장에서 종류가 「연차」이고 사용일자가 산정기간 안에 있는 건만 더합니다. 경조·병가는 빠집니다.')
lv.cell(row=nrow + 2, column=1,
        value='· 산정기간이 끝나면 잔여일수는 소멸하며, 소멸분에 대해 미사용 수당을 지급해야 합니다. (취업규칙 제23조제4항)')
for i in range(3):
    lv.cell(row=nrow + i, column=1).font = Font(name=F, size=9, color=GRAY)

# ============================================================
# 4. 휴가대장
# ============================================================
lg = wb.create_sheet('휴가대장')
lg.sheet_view.showGridLines = False
title_block(lg, '휴가대장', '휴가를 쓸 때마다 한 줄씩 기록합니다. 종류가 「연차」인 건만 연차현황의 사용일수에 반영됩니다.')
legend(lg, 3, '반차는 일수 칸에 0.5를 넣으세요.  ·  연속 휴가는 날짜별로 한 줄씩 나눠 적는 것이 집계에 정확합니다.')
hdr = ['사용일자', '사번', '성명', '휴가종류', '일수', '사유', '승인자', '신청일', '비고']
# 열 순서: A 사용일자 / B 사번 / C 성명 / D ... → SUMIFS 참조와 맞추기 위해 재배치
hdr = ['순번', '사번', '성명', '사용일자', '휴가종류', '일수', '사유', '승인자', '신청일']
header_row(lg, 4, hdr, [8, 10, 11, 13, 15, 8, 30, 11, 13], auto_cols=(1, 3))
lg.freeze_panes = 'D5'

for r in range(LV_FIRST, LV_LAST + 1):
    body_cell(lg, r, 1, f'=IF(B{r}="","",COUNTA($B${LV_FIRST}:$B{r}))', fmt='0', auto=True)
    body_cell(lg, r, 2)
    body_cell(lg, r, 3,
        f'=IFERROR(IF(B{r}="","",INDEX(직원명부!$B${EMP_FIRST}:$B${EMP_LAST},'
        f'MATCH(B{r},직원명부!$A${EMP_FIRST}:$A${EMP_LAST},0))),"사번 확인")', auto=True)
    body_cell(lg, r, 4, fmt='yyyy-mm-dd')
    body_cell(lg, r, 5)
    body_cell(lg, r, 6, fmt='0.0')
    body_cell(lg, r, 7, center=False)
    body_cell(lg, r, 8)
    body_cell(lg, r, 9, fmt='yyyy-mm-dd')

import datetime
exl = [None, 'M001', None, datetime.date(2026, 5, 4), '연차', 1, '개인 사유', '대표', datetime.date(2026, 4, 29)]
for i, v in enumerate(exl, start=1):
    if i in (1, 3) or v is None:
        continue
    cell = lg.cell(row=LV_FIRST, column=i, value=v)
    cell.font = Font(name=F, size=9, color=GRAY, italic=True)

dv(lg, 'E', LV_FIRST, LV_LAST, 'D', 9)

# ============================================================
# 5. 근태기록
# ============================================================
at = wb.create_sheet('근태기록')
at.sheet_view.showGridLines = False
title_block(at, '근태기록', '출퇴근 시각을 매일 기록합니다. 근로기준법 제42조에 따라 3년간 보존해야 합니다.')
legend(at, 3, '시각은 09:00 형식으로 입력합니다.  ·  휴게시간은 시간 단위 숫자(1 = 1시간)로 넣습니다.  ·  '
              '자정을 넘겨 퇴근한 경우에도 퇴근 시각만 그대로 적으면 근무시간이 맞게 계산됩니다.')
hdr = ['날짜', '요일', '사번', '성명', '출근', '퇴근', '휴게(h)', '근무시간', '연장시간',
       '야간시간', '보상휴가', '근태구분', '프로젝트', '비고']
header_row(at, 4, hdr, [13, 7, 10, 11, 9, 9, 10, 11, 11, 11, 11, 12, 18, 22],
           auto_cols=(2, 4, 8, 9, 10, 11))
at.freeze_panes = 'E5'

for r in range(AT_FIRST, AT_LAST + 1):
    body_cell(at, r, 1, fmt='yyyy-mm-dd')
    body_cell(at, r, 2,
        f'=IF(A{r}="","",CHOOSE(WEEKDAY(A{r},2),"월","화","수","목","금","토","일"))', auto=True)
    body_cell(at, r, 3)
    body_cell(at, r, 4,
        f'=IFERROR(IF(C{r}="","",INDEX(직원명부!$B${EMP_FIRST}:$B${EMP_LAST},'
        f'MATCH(C{r},직원명부!$A${EMP_FIRST}:$A${EMP_LAST},0))),"사번 확인")', auto=True)
    body_cell(at, r, 5, fmt='hh:mm')
    body_cell(at, r, 6, fmt='hh:mm')
    body_cell(at, r, 7, fmt='0.0')
    body_cell(at, r, 8,
        f'=IF(OR(E{r}="",F{r}=""),"",ROUND(MOD(F{r}-E{r},1)*24-N(G{r}),2))', fmt='0.00', auto=True)
    body_cell(at, r, 9, f'=IF(H{r}="","",ROUND(MAX(0,H{r}-8),2))', fmt='0.00', auto=True)
    # 야간근로 22:00~06:00. 09시 출근을 전제로 퇴근 시각만으로 산정한다.
    body_cell(at, r, 10,
        f'=IF(OR(E{r}="",F{r}=""),"",ROUND(IF(MOD(F{r},1)>=TIME(22,0,0),'
        f'(MOD(F{r},1)-TIME(22,0,0))*24,IF(MOD(F{r},1)<TIME(6,0,0),'
        f'(MOD(F{r},1)+1-TIME(22,0,0))*24,0)),2))', fmt='0.00', auto=True)
    # 보상휴가 = 가산율을 포함한 환산 시간 (취업규칙 제21조제4항)
    body_cell(at, r, 11,
        f'=IF(H{r}="","",ROUND(IF(L{r}="휴일근로",MIN(H{r},8)*1.5+MAX(0,H{r}-8)*2,'
        f'I{r}*1.5)+N(J{r})*0.5,2))', fmt='0.00', auto=True)
    body_cell(at, r, 12)                     # 근태구분
    body_cell(at, r, 13)                     # 프로젝트
    body_cell(at, r, 14, center=False)       # 비고

exa = [datetime.date(2026, 5, 6), None, 'M001', None,
       datetime.time(9, 0), datetime.time(23, 0), 1, None, None, None, None,
       '정상', '코오롱 헌인마을', '마감 대응']
for i, v in enumerate(exa, start=1):
    if i in (2, 4, 8, 9, 10, 11) or v is None:
        continue
    cell = at.cell(row=AT_FIRST, column=i, value=v)
    cell.font = Font(name=F, size=9, color=GRAY, italic=True)
    if i in (5, 6):
        cell.number_format = 'hh:mm'
    if i == 1:
        cell.number_format = 'yyyy-mm-dd'

dv(at, 'L', AT_FIRST, AT_LAST, 'C', 9)

nrow = AT_LAST + 2
at.cell(row=nrow, column=1,
        value='· 「연장시간」은 1일 8시간 초과분입니다. 이 시간에는 통상임금의 50%를 가산해 지급합니다. '
              '휴일근로는 8시간 이내 50%, 8시간 초과분은 100%입니다. (취업규칙 제21조제2항)')
at.cell(row=nrow + 1, column=1,
        value='· 5인 미만 사업장에는 가산수당 지급의무가 법으로 강제되지 않지만, 마인드는 근로계약과 통일하기 위해 '
              '취업규칙으로 정해 지급합니다. 5명 이상이 되면 여기에 주 12시간 연장 한도가 더해집니다.')
at.cell(row=nrow + 2, column=1,
        value='· 「프로젝트」 칸에 그날 주로 작업한 프로젝트명을 적으면 프로젝트집계 시트에서 투입시간이 자동으로 모입니다. '
              '프로젝트명은 매번 같은 표기를 쓰세요. 띄어쓰기가 다르면 다른 프로젝트로 집계됩니다.')
at.cell(row=nrow + 3, column=1,
        value='· 「야간시간」은 22:00~06:00 사이의 근로시간이며 퇴근 시각으로 자동 산정합니다. 자정을 넘겨도 계산됩니다. '
              '「보상휴가」는 가산율을 반영한 환산 시간입니다 — 연장 × 1.5 + 야간 × 0.5, 휴일근로는 8시간 이내 × 1.5, '
              '초과분 × 2.0. (취업규칙 제21조제4항)')
at.cell(row=nrow + 4, column=1,
        value='· 보상휴가는 1대1이 아닙니다. 23:00까지 근무하면 8시간(하루치)이 쌓입니다. 사용과 잔여는 보상휴가 시트에서 봅니다.')
for i in range(5):
    at.cell(row=nrow + i, column=1).font = Font(name=F, size=9, color=GRAY)

# ============================================================
# 6. 월별집계
# ============================================================
sm = wb.create_sheet('월별집계')
sm.sheet_view.showGridLines = False
title_block(sm, '월별집계', '조회할 달의 1일 날짜만 바꾸면 그 달 실적이 직원별로 자동 집계됩니다.')
sm['A3'] = '조회 월'
sm['A3'].font = Font(name=F, size=9, bold=True, color=INK)
sm['B3'] = datetime.date(2026, 5, 1)
sm['B3'].font = Font(name=F, size=10, bold=True, color=BLUE)
sm['B3'].number_format = 'yyyy"년" mm"월"'
sm['B3'].fill = PatternFill('solid', fgColor=YEL)
sm['B3'].border = BOX
sm['C3'] = '← 반드시 해당 월 1일 날짜로 넣으세요. (예: 2026-05-01)'
sm['C3'].font = Font(name=F, size=9, color=GRAY)

hdr = ['사번', '성명', '재직상태', '근무일수', '총 근무시간', '연장시간',
       '지각', '조퇴', '결근', '연차사용', '기타휴가']
header_row(sm, 4, hdr, [10, 11, 11, 11, 13, 11, 8, 8, 8, 11, 11],
           auto_cols=tuple(range(1, 12)))
sm.freeze_panes = 'C5'

A_ = f'근태기록!$A${AT_FIRST}:$A${AT_LAST}'
C_ = f'근태기록!$C${AT_FIRST}:$C${AT_LAST}'
H_ = f'근태기록!$H${AT_FIRST}:$H${AT_LAST}'
I_ = f'근태기록!$I${AT_FIRST}:$I${AT_LAST}'
J_ = f'근태기록!$L${AT_FIRST}:$L${AT_LAST}'
LB = f'휴가대장!$B${LV_FIRST}:$B${LV_LAST}'
LD = f'휴가대장!$D${LV_FIRST}:$D${LV_LAST}'
LE = f'휴가대장!$E${LV_FIRST}:$E${LV_LAST}'
LF = f'휴가대장!$F${LV_FIRST}:$F${LV_LAST}'
RANGE = f'{A_},">="&$B$3,{A_},"<="&EOMONTH($B$3,0)'
LRANGE = f'{LD},">="&$B$3,{LD},"<="&EOMONTH($B$3,0)'

for r in range(EMP_FIRST, EMP_LAST + 1):
    body_cell(sm, r, 1, f'=IF(직원명부!A{r}="","",직원명부!A{r})', auto=True)
    body_cell(sm, r, 2, f'=IF(직원명부!B{r}="","",직원명부!B{r})', auto=True)
    body_cell(sm, r, 3, f'=IF(직원명부!F{r}="","",직원명부!F{r})', auto=True)
    g = f'IF($A{r}="","",'
    body_cell(sm, r, 4, f'={g}COUNTIFS({C_},$A{r},{RANGE},{H_},">0"))', fmt='0', auto=True)
    body_cell(sm, r, 5, f'={g}SUMIFS({H_},{C_},$A{r},{RANGE}))', fmt='0.00', auto=True)
    body_cell(sm, r, 6, f'={g}SUMIFS({I_},{C_},$A{r},{RANGE}))', fmt='0.00', auto=True)
    body_cell(sm, r, 7, f'={g}COUNTIFS({C_},$A{r},{RANGE},{J_},"지각"))', fmt='0', auto=True)
    body_cell(sm, r, 8, f'={g}COUNTIFS({C_},$A{r},{RANGE},{J_},"조퇴"))', fmt='0', auto=True)
    body_cell(sm, r, 9, f'={g}COUNTIFS({C_},$A{r},{RANGE},{J_},"결근"))', fmt='0', auto=True)
    body_cell(sm, r, 10, f'={g}SUMIFS({LF},{LB},$A{r},{LE},"연차",{LRANGE}))', fmt='0.0', auto=True)
    body_cell(sm, r, 11, f'={g}SUMIFS({LF},{LB},$A{r},{LRANGE})-'
                         f'SUMIFS({LF},{LB},$A{r},{LE},"연차",{LRANGE}))', fmt='0.0', auto=True)

nrow = EMP_LAST + 2
sm.cell(row=nrow, column=1,
        value='· 「근무일수」는 근태기록에 출근·퇴근이 모두 적혀 근무시간이 0보다 큰 날의 수입니다.')
sm.cell(row=nrow + 1, column=1,
        value='· 「기타휴가」는 연차를 뺀 나머지(경조·병가·공가·출산·육아 등) 전부의 합계입니다.')
for i in range(2):
    sm.cell(row=nrow + i, column=1).font = Font(name=F, size=9, color=GRAY)

# ============================================================
# 7. 프로젝트집계
# ============================================================
pj = wb.create_sheet('프로젝트집계')
pj.sheet_view.showGridLines = False
title_block(pj, '프로젝트집계', '근태기록의 「프로젝트」 칸을 모아 투입시간을 집계합니다. 다음 견적의 근거가 됩니다.')
pj['A3'] = '조회 기간'
pj['A3'].font = Font(name=F, size=9, bold=True, color=INK)
for col, val in (('B', datetime.date(2026, 1, 1)), ('C', datetime.date(2026, 12, 31))):
    c = pj[col + '3']
    c.value = val
    c.font = Font(name=F, size=10, bold=True, color=BLUE)
    c.number_format = 'yyyy-mm-dd'
    c.fill = PatternFill('solid', fgColor=YEL)
    c.border = BOX
    c.alignment = Alignment(horizontal='center')
pj['D3'] = '← 시작일과 종료일. 이 기간에 기록된 근무시간만 집계합니다.'
pj['D3'].font = Font(name=F, size=9, color=GRAY)

PJ_FIRST, PJ_LAST = 5, 34
header_row(pj, 4, ['프로젝트', '총 근무시간', '연장시간', '근무일수', '인건비 환산', '비고'],
           [26, 13, 12, 11, 15, 30], auto_cols=(2, 3, 4, 5))
pj.freeze_panes = 'B5'

AK = f'근태기록!$M${AT_FIRST}:$M${AT_LAST}'
AA = f'근태기록!$A${AT_FIRST}:$A${AT_LAST}'
AH = f'근태기록!$H${AT_FIRST}:$H${AT_LAST}'
AI = f'근태기록!$I${AT_FIRST}:$I${AT_LAST}'
PER = f'{AA},">="&$B$3,{AA},"<="&$C$3'

for r in range(PJ_FIRST, PJ_LAST + 1):
    body_cell(pj, r, 1, center=False)
    body_cell(pj, r, 2, f'=IF($A{r}="","",SUMIFS({AH},{AK},$A{r},{PER}))',
              fmt='0.00', auto=True)
    body_cell(pj, r, 3, f'=IF($A{r}="","",SUMIFS({AI},{AK},$A{r},{PER}))',
              fmt='0.00', auto=True)
    body_cell(pj, r, 4, f'=IF($A{r}="","",COUNTIFS({AK},$A{r},{PER},{AH},">0"))',
              fmt='0', auto=True)
    body_cell(pj, r, 5, f'=IF(OR($A{r}="",$B$36=""),"",ROUND(B{r}*$B$36,0))',
              fmt='#,##0', auto=True)
    body_cell(pj, r, 6, center=False)

pj['A36'] = '시간당 단가 (원)'
pj['A36'].font = Font(name=F, size=9, bold=True, color=INK)
pj['B36'] = '=IFERROR(ROUND(직원명부!I5/209,0),"")'
pj['B36'].font = Font(name=F, size=10, bold=True, color=INK)
pj['B36'].number_format = '#,##0'
pj['B36'].fill = PatternFill('solid', fgColor=SOFT_FILL)
pj['B36'].border = BOX
pj['B36'].alignment = Alignment(horizontal='center')
pj['C36'] = '← 직원명부 첫 사원의 월 지급액 ÷ 209시간. 인원이 여럿이면 평균 단가를 직접 넣으세요.'
pj['C36'].font = Font(name=F, size=9, color=GRAY)

for i, t in enumerate([
    '· 프로젝트명은 근태기록의 「프로젝트」 칸에 적은 것과 글자가 정확히 같아야 집계됩니다.',
    '· 「인건비 환산」은 원가 감각을 잡기 위한 값입니다. 4대보험 회사부담분과 간접비가 빠져 있어 '
    '실제 원가는 이보다 높습니다. 견적에 그대로 쓰지 마세요.',
    '· 프로젝트가 끝나면 총 근무시간을 견적서와 대조해 두세요. 다음 견적의 근거가 됩니다.',
], start=38):
    pj.cell(row=i, column=1, value=t).font = Font(name=F, size=9, color=GRAY)

# ============================================================
# 8. 보상휴가
# ============================================================
cp = wb.create_sheet('보상휴가')
cp.sheet_view.showGridLines = False
title_block(cp, '보상휴가', '연장·야간·휴일근로 수당을 갈음하는 휴가입니다. 1대1이 아니라 가산율을 반영해 환산합니다.')
cp['A3'] = '조회 기간'
cp['A3'].font = Font(name=F, size=9, bold=True, color=INK)
for col, val in (('B', datetime.date(2026, 1, 1)), ('C', datetime.date(2026, 12, 31))):
    c = cp[col + '3']
    c.value = val
    c.font = Font(name=F, size=10, bold=True, color=BLUE)
    c.number_format = 'yyyy-mm-dd'
    c.fill = PatternFill('solid', fgColor=YEL)
    c.border = BOX
    c.alignment = Alignment(horizontal='center')
cp['D3'] = '← 이 기간에 발생하고 사용한 보상휴가만 집계합니다.'
cp['D3'].font = Font(name=F, size=9, color=GRAY)

CP_FIRST, CP_LAST = 5, 14
header_row(cp, 4, ['사번', '성명', '발생 시간', '사용 시간', '잔여 시간', '잔여 일수', '비고'],
           [10, 12, 13, 13, 13, 12, 30], auto_cols=(1, 2, 3, 4, 5, 6))
cp.freeze_panes = 'C5'

CK = f'근태기록!$K${AT_FIRST}:$K${AT_LAST}'
CC = f'근태기록!$C${AT_FIRST}:$C${AT_LAST}'
CA = f'근태기록!$A${AT_FIRST}:$A${AT_LAST}'
UF, UL = 19, 118          # 사용 기록 행
UD = f'$A${UF}:$A${UL}'
US = f'$B${UF}:$B${UL}'
UH = f'$D${UF}:$D${UL}'

for r in range(CP_FIRST, CP_LAST + 1):
    e = 'EMPROW'
    er = EMP_FIRST + (r - CP_FIRST)
    body_cell(cp, r, 1, f'=IF(직원명부!A{er}="","",직원명부!A{er})', auto=True)
    body_cell(cp, r, 2, f'=IF(직원명부!B{er}="","",직원명부!B{er})', auto=True)
    body_cell(cp, r, 3,
        f'=IF($A{r}="","",SUMIFS({CK},{CC},$A{r},{CA},">="&$B$3,{CA},"<="&$C$3))',
        fmt='0.00', auto=True)
    body_cell(cp, r, 4,
        f'=IF($A{r}="","",SUMIFS({UH},{US},$A{r},{UD},">="&$B$3,{UD},"<="&$C$3))',
        fmt='0.00', auto=True)
    c = body_cell(cp, r, 5, f'=IF($A{r}="","",C{r}-D{r})', fmt='0.00', auto=True)
    c.font = Font(name=F, size=9, bold=True, color=INK)
    body_cell(cp, r, 6, f'=IF($A{r}="","",ROUND(E{r}/8,2))', fmt='0.00', auto=True)
    body_cell(cp, r, 7, center=False)

cp['A17'] = '사용 기록'
cp['A17'].font = Font(name=F, size=11, bold=True, color=INK)
cp['C17'] = '보상휴가를 쓸 때마다 한 줄씩 적습니다. 1시간 단위로 사용할 수 있습니다.'
cp['C17'].font = Font(name=F, size=9, color=GRAY)
header_row(cp, 18, ['사용일자', '사번', '성명', '사용 시간', '승인자', '비고', ''],
           [13, 10, 12, 13, 12, 26, 4], auto_cols=(3,))
for r in range(UF, UL + 1):
    body_cell(cp, r, 1, fmt='yyyy-mm-dd')
    body_cell(cp, r, 2)
    body_cell(cp, r, 3,
        f'=IFERROR(IF(B{r}="","",INDEX(직원명부!$B${EMP_FIRST}:$B${EMP_LAST},'
        f'MATCH(B{r},직원명부!$A${EMP_FIRST}:$A${EMP_LAST},0))),"사번 확인")', auto=True)
    body_cell(cp, r, 4, fmt='0.0')
    body_cell(cp, r, 5)
    body_cell(cp, r, 6, center=False)

for i, t in enumerate([
    '· 발생 시간은 근태기록의 「보상휴가」 열을 모은 값입니다. 연장 × 1.5 + 야간 × 0.5 로 환산되어 있습니다.',
    '· 23:00까지 근무하면 8시간이 쌓입니다. 하루 휴무와 같습니다.',
    '· 사유가 발생한 달의 말일부터 3개월 안에 써야 합니다. 남으면 소멸하지 않고 수당으로 지급해야 합니다. '
    '(취업규칙 제21조제5항)',
    '· 사원이 보상휴가를 원하지 않으면 수당으로 지급합니다. 회사가 일방적으로 휴가로 갈음할 수 없습니다. (제21조제7항)',
    '· 운영하려면 근로자대표와의 서면합의가 필요합니다. 「보상휴가제 서면합의서」를 먼저 체결하세요.',
], start=UL + 2):
    cp.cell(row=i, column=1, value=t).font = Font(name=F, size=9, color=GRAY)

wb.move_sheet('코드표', offset=7)
wb.calculation.fullCalcOnLoad = True   # 파일을 열 때 엑셀이 전체 수식을 다시 계산하도록
import os
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '마인드_근태연차_관리대장.xlsx')
wb.save(OUT)
print('saved', [s.title for s in wb.worksheets])
